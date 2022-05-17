import os
import easygui
from openpyxl.cell.cell import MergedCell
from openpyxl import load_workbook
from enum import Enum


class ObjectType(Enum):
    VAR = 7
    ARRAY = 8
    RECORD = 9


class DataType(Enum):
    BOOLEAN = 0
    INTEGER8 = 1
    INTEGER16 = 2
    INTEGER32 = 3
    INTEGER64 = 4
    UNSIGNED8 = 5
    UNSIGNED16 = 6
    UNSIGNED32 = 7
    UNSIGNED64 = 8
    REAL32 = 9
    REAL64 = 10
    VISIBLE_STRING = 11
    OCTET_STRING = 12
    UNICODE_STRING = 13
    DOMAIN = 14


class dictOb():

    def __init__(self, header_dict):
        self.primaryDict = header_dict.copy()
        self.subindex = []

    @property
    def name(self):
        return self.primaryDict.get("Name", None)

    @property
    def index(self):
        return self.primaryDict.get("Index", None)

    @property
    def DataType(self):
        return self.primaryDict.get("Data Type", None)

    @property
    def ObjectType(self):
        if len(self.subindex) > 1:
            return self.DataType
        else:
            return "VAR"






EDS_object_format = {
    "SubNumber": None,
    "ObjectType": None,
    "ParameterName": None
}

EDS_sub_format = {
    "ParameterName": None,
    "ParameterValue": None,
    "ObjectType": None,
    "DataType": None,
    "AccessType": None,
    "LowLimit": None,
    "HighLimit": None,
    "PDOMapping": None,
    "DefaultValue": None
}






def write_esd_object(object: dictOb) -> str:
    string = ""
    if len(object.subindex) > 1:
        string += "[{}]\n".format(hex(object.index)[2:])
        string += "{}={}\n".format("ParameterName", object.name)
        string += "{}={}\n".format("ObjectType", ObjectType[object.DataType.upper().replace(" ", "_")].value)
        string += ";StorageLocation=RAM\n"
        string += "{}={}\n\n".format("SubNumber", len(object.subindex))

    for item in object.subindex:
        if len(object.subindex) == 1:
            string += "[{}]\n".format(hex(object.index)[2:])
        else:
            string += "[{}sub{}]\n".format(hex(object.index)[2:], item["Sub-Index"])
        string += "{}={}\n".format("ParameterName", item["Name"])
        string += "{}={}\n".format("ObjectType", "0x7")
        string += ";StorageLocation=RAM\n"
        string += "{}={}\n".format("DataType", hex(DataType[item["Data Type"].upper().replace(" ", "_")].value))
        string += "{}={}\n".format("AccessType", item["Access Type"].lower())
        string += "{}={}\n".format("HighLimit", item["High Limit"])
        string += "{}={}\n".format("LowLimit", item["Low Limit"])
        value = item.get("Value", "")
        if isinstance(value, int):
            value = hex(value)
        string += "{}={}\n".format("DefaultValue", value)
        pdo = 0
        if item["Map to PDO?"] == "Yes":
            pdo = 1
        string += "{}={}\n\n".format("PDOMapping", pdo)

    return string

def fix_value(this_value):
    # Return if None
    if this_value is None:
        return this_value
    elif this_value == "None":
        return None
    # Return if not a string
    elif not isinstance(this_value, str):
        return this_value
    # Check if string is numberic
    elif this_value.isnumeric():
        return int(this_value)
    # Check if string is hex
    elif this_value[-1:] == 'h':
        try:
            int("0x"+this_value[:-1], 0)
        except:
            return this_value
        return int("0x"+this_value[:-1], 0)
    else:
        return this_value


file1 = easygui.fileopenbox("select test point file", default="C:/Repos/E_Moto/Firmware/SevconController/")

wb1 = load_workbook(file1)

wb1.active = wb1['Object Dictionary']

header_dict = {}
header_row = wb1.active[4]
for i in range(1, 18):
    header_dict[str(header_row[i].value)] = None

newOb = None
index = 0

Object_Dictionary = {}
for row in wb1.active.iter_rows(min_row=5):


    # Convert the row into a dictionary
    tempdict = header_dict.copy()
    for i, item in enumerate(tempdict):
        value = row[i+1].value
        value = fix_value(value)
        tempdict[item] = value

    # Check for single variable object
    if (tempdict["Index"] is not None) and (tempdict["Sub-Index"] == 0):
        newOb = dictOb(header_dict)
        newOb.primaryDict = tempdict.copy()
        newOb.subindex.append(tempdict.copy())

    # Check for array object
    elif tempdict["Sub-Index"] is None:
        newOb = dictOb(header_dict)
        newOb.primaryDict = tempdict.copy()
        index = 0

    # Check for sub-index objects
    elif isinstance(row[1], MergedCell):
        # always get the lastest "version" of an object
        if tempdict["Sub-Index"] != index:
            newOb.subindex.pop()
        tempdict["Index"] = newOb.primaryDict["Index"]
        newOb.subindex.append(tempdict.copy())
        index += 1

    # This should never happen
    else:
        print("Never get here")
        raise Exception("Unidentified Object at {}".format(row))

    if newOb.index =="None":
        print('wtf')
        raise Exception("Unidentified Object at {}".format(row))

    Object_Dictionary[newOb.index] = newOb
    print(newOb.index)

print("done")

# Fix objects with missing datatypes


path = os.path.dirname(file1)

with open(os.path.join(path, 'new_EDS.eds'), 'w') as f:

    Mandatory_object_area = [0x1000, 0x1001, 0x1018]
    Communication_object_area = range(0x1000, 0x1FFF)
    Manufacturer_object_area = range(0x2000, 0x5FFF)
    Device_object_area = range(0x6000, 0x9FFF)
    rpdo_range = range(0x1400, 0x16FF)
    tpdo_range = range(0x1800, 0x1AFF)

    range_comm = {}
    range_mfg = {}
    range_dev = {}
    ran_man = {}



    for object in Object_Dictionary:
        if Object_Dictionary[object].index in rpdo_range or Object_Dictionary[object].index in tpdo_range:
            continue
        if Object_Dictionary[object].index in Mandatory_object_area:
            ran_man[object] = Object_Dictionary[object]
        elif Object_Dictionary[object].index in Communication_object_area or Object_Dictionary[object].index in Device_object_area:
            range_comm[object] = Object_Dictionary[object]
        elif Object_Dictionary[object].index in Manufacturer_object_area:
            range_mfg[object] = Object_Dictionary[object]
        # elif Object_Dictionary[object].index in Device_object_area:
        #     range_dev[object] = Object_Dictionary[object]
        else:
            print('wtf')
            raise Exception("Out of range Object at {}".format(object))

    f.write("[MandatoryObjects]\nSupportedObjects={}\n".format(len(ran_man)))
    for num, object in enumerate(ran_man):
        f.write("{}={}\n".format(num+1, hex(object).upper()))
    f.write("\n")
    for object in ran_man:
        f.write(write_esd_object(Object_Dictionary[object]))

    f.write("[OptionObjects]\nSupportedObjects={}\n".format(len(range_comm)))
    for num, object in enumerate(range_comm):
        f.write("{}={}\n".format(num+1, hex(object).upper()))
    f.write("\n")
    for object in range_comm:
        f.write(write_esd_object(Object_Dictionary[object]))

    f.write("[ManufacturerObjects]\nSupportedObjects={}\n".format(len(range_mfg)))
    for num, object in enumerate(range_mfg):
        f.write("{}={}\n".format(num+1, hex(object).upper()))
    f.write("\n")
    for object in range_mfg:
        f.write(write_esd_object(Object_Dictionary[object]))

    # f.write("[OptionObjects]\nSupportedObjects={}\n".format(len(range_comm)))
    # for num, object in enumerate(range_comm):
    #     f.write("{}={}".format(num+1, hex(object)))
    # for object in range_comm:
    #     write_esd_object(Object_Dictionary[object])



print("done")


