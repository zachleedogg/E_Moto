import pygetwindow as gw
import pyautogui
import keyboard
import easygui
from threading import Thread
from time import sleep
import openpyxl


class refDesSearch:

    window_flag = False
    key_flag = False
    quit = False
    des = 0
    bom_headers = []
    bom_list = []
    component_index = -1

    def __init__(self):
        # get BOM file by user input
        f = easygui.fileopenbox(default="C:\Repos\E_Moto\Hardware")
        # load the excel workbook
        wb = openpyxl.load_workbook(filename=f)
        ws = wb.active
        # get the header row information
        for item in ws[1]:
            self.bom_headers.append(item.value)
        # create a dictionary of all the BOM items
        for row in range(2, ws.max_row+1):
            this_dict = {}
            for num, param in enumerate(ws[row]):
                this_dict[self.bom_headers[num]] = str(param.value)
            self.bom_list.append(this_dict)

        self.bom_list = sorted(self.bom_list, key=lambda i: (i['MFG_PN'], i['RefDes']))

        ref_des_types = []
        for item in self.bom_list:
            type = self.get_ref_des_type(item["RefDes"])
            if type not in ref_des_types:
                ref_des_types.append(type)

        new_bom_list = []
        for red_des_type in sorted(ref_des_types):
            for item in self.bom_list:
                if item["RefDes"].startswith(red_des_type):
                    new_bom_list.append(item)

        self.bom_list = new_bom_list

    @staticmethod
    def get_ref_des_type(ref_des):
        ret_char = ''
        for char in ref_des:
            if char.isalpha():
                ret_char += char
            else:
                break
        return ret_char

    def show_window(self):
        while True:
            if self.quit:
                return
            sleep(.1)
            if self.window_flag is True:
                string = ""
                for item in self.bom_list[self.component_index]:
                    string += item + ": " + self.bom_list[self.component_index][item] + "\n"
                string += "\n\nUse < and > keys to cycle through components. Press Cancel or q to quit"
                if easygui.ccbox(string, 'BOM_ITEM') is False:
                    self.quit = True
                    return
                self.des += 1
                self.window_flag = False

    def close_window(self):
        win = gw.getWindowsWithTitle('BOM_ITEM')
        if len(win):
            win[0].activate()
            sleep(.01)
            pyautogui.press("esc")

    def keyboard_input(self):
        while True:
            if self.quit:
                return
            sleep(.01)
            if keyboard.is_pressed("."):
                self.component_index += 1
                self.key_flag = True
            elif keyboard.is_pressed(","):
                self.component_index -= 1
                self.key_flag = True
            elif keyboard.is_pressed("q"):
                self.quit = True
                return

            if self.key_flag:
                if self.component_index < 0:
                    self.component_index = 0
                if self.component_index >= len(self.bom_list):
                    self.component_index = len(self.bom_list) - 1
                self.do_the_search()

    def do_the_search(self):
        self.key_flag = False

        self.close_window()

        all_windows = gw.getAllWindows()
        for window in all_windows:
            if "PCB Editor" in window.title or "PcbNew" in window.title:
                window.activate()
                break

        pyautogui.hotkey('ctrl', 'f')

        pyautogui.typewrite(str(self.bom_list[self.component_index]["RefDes"]))

        pyautogui.press("enter")
        pyautogui.hotkey("shift", "tab")
        pyautogui.press("enter")
        self.window_flag = True


wind = refDesSearch()

t = Thread(target=wind.show_window)

t.start()

wind.keyboard_input()

wind.close_window()


